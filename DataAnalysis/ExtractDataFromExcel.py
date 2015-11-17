from openpyxl import load_workbook
from openpyxl import Workbook
import os

DATA_PATH = r'/Users/Dijkstraaaaa/Documents/1'
OUTPUT_PATH = r'/Users/Dijkstraaaaa/Documents/BS Data'

wb = load_workbook(filename=r'%s/1 2.xlsx' % DATA_PATH)
ws = wb.worksheets[0]  # ws is now an IterableWorksheet

# The wanted column in excel
column_dic = {"SUBNET_ID": 5, "NETELE_ID": 6, "START_TIME": 2, "END_TIME": 3, "CPU_PEAK_USAGE": 9,
              "CPU_AVARATE_USAGE": 10, "CPU_OVERLOAD_RATE": 11}
column_name = ["SUBNET_ID", "NETELE_ID", "START_TIME", "END_TIME", "CPU_PEAK_USAGE", "CPU_AVARATE_USAGE", "CPU_OVERLOAD_RATE"]

# The first 4 rows are table info
for row in range(ws.min_row + 4, ws.max_row + 1):
	print "Processing ROW number : %dn" % row

	subnet_id = ws.cell(row=row, column=column_dic['SUBNET_ID']).value
	netele_id = ws.cell(row=row, column=column_dic['NETELE_ID']).value
	file_path = r'%s/%s' % (OUTPUT_PATH, subnet_id) + "_" + '%s.xlsx' % netele_id

	if os.path.isfile(file_path):
		output_file = load_workbook(filename=file_path)
		# insert at the end (default)
		output_file_sheet = output_file.active
	else:
		output_file = Workbook()
		# insert at the end (default)
		output_file_sheet = output_file.active
		for i in range(len(column_name)):
			output_file_sheet.cell(row=1, column=i + 1).value = column_name[i]

	content_list = []
	for item in column_name:
		content_list.append(ws.cell(row=row, column=column_dic[item]).value)
	output_file_sheet.append(content_list)

	output_file.save(file_path)
