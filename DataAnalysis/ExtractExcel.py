from openpyxl import load_workbook
import time
import os

DATA_PATH = r'/Users/Dijkstraaaaa/Documents/1'
OUTPUT_PATH = r'/Users/Dijkstraaaaa/Documents/BS Data2'

wb = load_workbook(filename=r'%s/1 2.xlsx' % DATA_PATH)
ws = wb.worksheets[0]  # ws is now an IterableWorksheet

# The wanted column in excel
column_dic = {"SUBNET_ID": 5, "NETELE_ID": 6, "START_TIME": 2, "END_TIME": 3, "CPU_PEAK_USAGE": 9,
              "CPU_AVARATE_USAGE": 10, "CPU_OVERLOAD_RATE": 11}
column_name = ["SUBNET_ID", "NETELE_ID", "START_TIME", "END_TIME", "CPU_PEAK_USAGE", "CPU_AVARATE_USAGE",
               "CPU_OVERLOAD_RATE"]

# The first 4 rows are table info
for row in range(ws.min_row + 4, ws.max_row + 1):
	start = time.time()

	print "Processing ROW number : %dn" % row
	subnet_id = ws.cell(row=row, column=column_dic['SUBNET_ID']).value
	netele_id = ws.cell(row=row, column=column_dic['NETELE_ID']).value
	file_path = r'%s/%s' % (OUTPUT_PATH, subnet_id) + "_" + '%s.txt' % netele_id

	if not os.path.isfile(file_path):
		output_file = open(file_path, 'a')
		output_file.write("\t".join(column_name) + "\n")
	else:
		output_file = open(file_path, 'a')

	content_list = []
	for item in column_name:
		content_list.append(ws.cell(row=row, column=column_dic[item]).value)

	output_file.write("\t".join(str(v) for v in content_list) + "\n")

	output_file.close()

	end = time.time()
	print end - start
