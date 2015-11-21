# -*- coding: UTF-8 -*-
import os
import csv
from openpyxl import load_workbook


def generate_file_list(day_num, _file_path):
    # use DAY_NUM to generate empty 2-D list
    _file_matrix = []
    while day_num != 0:
        _file_matrix.append([])
        day_num -= 1

    # walk the file path
    for root, dirs, files in os.walk(_file_path):
        for f in files:
            if f.startswith("TYPE3"):
                f_words = f.split("_")
                # date index
                day_info = f_words[-2]
                day_num = int(day_info[3:])
                # file index
                serial_info = f_words[-1]
                _file_matrix[day_num - 1].append(serial_info)
    return _file_matrix


def extract_csv_file(_csvfile_path, _output_folder_path):
    column_dic = {"SUBNET_ID": 4, "NETELE_ID": 5, "START_TIME": 1, "END_TIME": 2, "CPU_PEAK_USAGE": 8,
                  "CPU_AVERAGE_USAGE": 9, "CPU_OVERLOAD_RATE": 10, "MEMORY_PEAK_USAGE": 11, "MEMORY_AVERAGE_USAGE": 12,
                  "MEMORY_OVERLOAD_RATE": 13}
    column_name = ["SUBNET_ID", "NETELE_ID", "START_TIME", "END_TIME", "CPU_PEAK_USAGE", "CPU_AVERAGE_USAGE",
                   "CPU_OVERLOAD_RATE", "MEMORY_PEAK_USAGE", "MEMORY_AVERAGE_USAGE",
                   "MEMORY_OVERLOAD_RATE"]
    with open(_csvfile_path, 'rb') as csvfile:
        filereader = csv.reader(csvfile)
        for row in filereader:
            subnet_id = row[column_dic['SUBNET_ID']]
            netele_id = row[column_dic['NETELE_ID']]
            _output_file_name = '_'.join([str(subnet_id), str(netele_id)]) + '.txt'
            _output_file_path = os.path.join(_output_folder_path, _output_file_name)
            if not os.path.isfile(_output_file_path):
                output_file = open(_output_file_path, 'a')
                output_file.write("\t".join(column_name) + "\n")
            else:
                output_file = open(_output_file_path, 'a')
            content_list = []
            for item in column_name:
                content_list.append(row[column_dic[item]])
            output_file.write("\t".join(str(v) for v in content_list) + "\n")
            output_file.close()


def extract_excel_file(_excel_file_path, _output_folder_path):
    _work_book = load_workbook(filename=_excel_file_path)
    _work_sheet = _work_book.worksheets[0]  # ws is now an IterableWorksheet
    # The wanted column in excel
    column_dic = {"SUBNET_ID": 5, "NETELE_ID": 6, "START_TIME": 2, "END_TIME": 3, "CPU_PEAK_USAGE": 9,
                  "CPU_AVERAGE_USAGE": 10, "CPU_OVERLOAD_RATE": 11, "MEMORY_PEAK_USAGE": 12, "MEMORY_AVERAGE_USAGE": 13,
                  "MEMORY_OVERLOAD_RATE": 14}
    column_name = ["SUBNET_ID", "NETELE_ID", "START_TIME", "END_TIME", "CPU_PEAK_USAGE", "CPU_AVERAGE_USAGE",
                   "CPU_OVERLOAD_RATE", "MEMORY_PEAK_USAGE", "MEMORY_AVERAGE_USAGE", "MEMORY_OVERLOAD_RATE"]
    # The first 4 rows are table info
    for row in range(_work_sheet.min_row + 4, _work_sheet.max_row + 1):
        subnet_id = _work_sheet.cell(row=row, column=column_dic['SUBNET_ID']).value
        netele_id = _work_sheet.cell(row=row, column=column_dic['NETELE_ID']).value
        _output_file_name = '_'.join([str(subnet_id), str(netele_id)]) + '.txt'
        _output_file_path = os.path.join(_output_folder_path, _output_file_name)
        if not os.path.isfile(_output_file_path):
            output_file = open(_output_file_path, 'a')
            output_file.write("\t".join(column_name) + "\n")
        else:
            output_file = open(_output_file_path, 'a')
        content_list = []
        for item in column_name:
            content_list.append(_work_sheet.cell(row=row, column=column_dic[item]).value)
        output_file.write("\t".join(str(v) for v in content_list) + "\n")
        output_file.close()


DATA_PATH = r'/Users/Dijkstraaaaa/Documents/Data/Mon1'
OUTPUT_PATH = r'/Users/Dijkstraaaaa/Documents/BS Data'

file_matrix = generate_file_list(31, DATA_PATH)
for i in range(len(file_matrix)):
    for j in range(len(file_matrix[i])):
        file_name = 'TYPE3_Mon1_Day%s' % (i + 1) + '_%s' % file_matrix[i][j]
        print file_name
        date = 'Day' + str(i + 1)
        file_path = os.path.join(DATA_PATH, date, 'TYPE3', file_name)
        if file_path.endswith("xlsx"):
            extract_excel_file(file_path, OUTPUT_PATH)
        if file_path.endswith("csv"):
            extract_csv_file(file_path, OUTPUT_PATH)
